from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import FIELD_ALIASES
from .security import extract_zip_safely


@dataclass
class ProcessResult:
    output_path: Path
    stats: dict[str, int]
    file_logs: list[dict[str, Any]] = field(default_factory=list)


def _text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def _normalize_key(value: Any) -> str:
    text = _text(value).lower()
    return re.sub(r"[\s_\-—－:：()（）\[\]【】/\\]+", "", text)


_ALIAS_LOOKUP: dict[str, str] = {}
for canonical, aliases in FIELD_ALIASES.items():
    for alias in aliases | {canonical}:
        _ALIAS_LOOKUP[_normalize_key(alias)] = canonical


def canonical_column(name: Any) -> str:
    raw = _text(name)
    normalized = _normalize_key(raw)
    return _ALIAS_LOOKUP.get(normalized, raw or "未命名列")


def _dedupe_column_names(columns: Iterable[Any]) -> list[str]:
    counts: dict[str, int] = {}
    output: list[str] = []
    for col in columns:
        base = _text(col) or "未命名列"
        counts[base] = counts.get(base, 0) + 1
        output.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return output


def _header_score(row: pd.Series) -> float:
    values = [_text(v) for v in row.tolist()]
    non_empty = [v for v in values if v]
    if len(non_empty) < 2:
        return -1
    known = sum(1 for v in non_empty if _normalize_key(v) in _ALIAS_LOOKUP)
    unique_ratio = len(set(non_empty)) / max(1, len(non_empty))
    long_penalty = sum(1 for v in non_empty if len(v) > 30)
    numeric_penalty = sum(1 for v in non_empty if re.fullmatch(r"\d+(\.\d+)?", v or ""))
    return len(non_empty) + known * 5 + unique_ratio - long_penalty * 1.5 - numeric_penalty * 0.5


def _read_csv(path: Path, **kwargs: Any) -> pd.DataFrame:
    errors: list[Exception] = []
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return pd.read_csv(path, encoding=encoding, **kwargs)
        except UnicodeDecodeError as exc:
            errors.append(exc)
    raise errors[-1]


def _read_raw(path: Path, sheet_name: str | int | None = None, nrows: int | None = None) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return _read_csv(path, header=None, nrows=nrows, dtype=object)
    return pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=nrows, dtype=object)


def _read_with_header(path: Path, header_row: int, sheet_name: str | int | None = None) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return _read_csv(path, header=header_row, dtype=object)
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row, dtype=object)


def detect_header_row(path: Path, sheet_name: str | int | None = None) -> int:
    raw = _read_raw(path, sheet_name=sheet_name, nrows=30)
    if raw.empty:
        return 0
    scores = [(_header_score(raw.iloc[index]), index) for index in range(len(raw))]
    score, index = max(scores, key=lambda item: item[0])
    return int(index if score >= 1 else 0)


def _sheet_names(path: Path) -> list[str | int | None]:
    if path.suffix.lower() == ".csv":
        return [None]
    with pd.ExcelFile(path) as workbook:
        return workbook.sheet_names


def _coalesce_duplicate_canonical_columns(frame: pd.DataFrame) -> pd.DataFrame:
    grouped: dict[str, list[str]] = {}
    for col in frame.columns:
        canonical = canonical_column(col)
        grouped.setdefault(canonical, []).append(col)

    result = pd.DataFrame(index=frame.index)
    for canonical, originals in grouped.items():
        series = frame[originals[0]]
        for extra in originals[1:]:
            series = series.where(series.map(_text) != "", frame[extra])
        result[canonical] = series
    return result


def _clean_frame(frame: pd.DataFrame, source_file: str, source_sheet: str) -> pd.DataFrame:
    frame = frame.copy()
    frame.columns = _dedupe_column_names(frame.columns)
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if frame.empty:
        return frame

    normalized_headers = [_normalize_key(col) for col in frame.columns]
    repeated_mask = frame.apply(
        lambda row: sum(
            _normalize_key(value) == normalized_headers[idx]
            for idx, value in enumerate(row.tolist()[: len(normalized_headers)])
        )
        >= max(2, len(normalized_headers) // 2),
        axis=1,
    )
    frame = frame.loc[~repeated_mask].copy()
    frame = _coalesce_duplicate_canonical_columns(frame)

    for col in frame.columns:
        frame[col] = frame[col].map(lambda value: _text(value))

    frame = frame.loc[frame.apply(lambda row: any(_text(v) for v in row), axis=1)].copy()
    if frame.empty:
        return frame

    frame["来源文件"] = source_file
    frame["来源工作表"] = source_sheet
    return frame.reset_index(drop=True)


def read_single_file(path: Path) -> tuple[list[pd.DataFrame], list[dict[str, Any]]]:
    frames: list[pd.DataFrame] = []
    logs: list[dict[str, Any]] = []

    for sheet in _sheet_names(path):
        sheet_label = "CSV" if sheet is None else str(sheet)
        try:
            header_row = detect_header_row(path, sheet)
            frame = _read_with_header(path, header_row, sheet)
            cleaned = _clean_frame(frame, path.name.split("_", 1)[-1], sheet_label)
            if cleaned.empty:
                logs.append(
                    {
                        "文件名": path.name.split("_", 1)[-1],
                        "工作表": sheet_label,
                        "状态": "跳过",
                        "有效行数": 0,
                        "说明": "未发现有效数据",
                    }
                )
                continue
            frames.append(cleaned)
            logs.append(
                {
                    "文件名": path.name.split("_", 1)[-1],
                    "工作表": sheet_label,
                    "状态": "成功",
                    "有效行数": len(cleaned),
                    "说明": f"识别表头第 {header_row + 1} 行",
                }
            )
        except Exception as exc:
            logs.append(
                {
                    "文件名": path.name.split("_", 1)[-1],
                    "工作表": sheet_label,
                    "状态": "失败",
                    "有效行数": 0,
                    "说明": str(exc)[:300],
                }
            )
    return frames, logs


def collect_excel_files(uploaded_paths: list[Path], working_dir: Path) -> tuple[list[Path], list[dict[str, Any]]]:
    excel_files: list[Path] = []
    logs: list[dict[str, Any]] = []
    zip_dir = working_dir / "unzipped"

    for path in uploaded_paths:
        if path.suffix.lower() == ".zip":
            try:
                items = extract_zip_safely(path, zip_dir / path.stem)
                excel_files.extend(items)
                if not items:
                    logs.append(
                        {
                            "文件名": path.name.split("_", 1)[-1],
                            "工作表": "-",
                            "状态": "跳过",
                            "有效行数": 0,
                            "说明": "ZIP 中没有支持的表格文件",
                        }
                    )
            except Exception as exc:
                logs.append(
                    {
                        "文件名": path.name.split("_", 1)[-1],
                        "工作表": "-",
                        "状态": "失败",
                        "有效行数": 0,
                        "说明": str(exc)[:300],
                    }
                )
        else:
            excel_files.append(path)
    return excel_files, logs


def _most_complete_rows(frame: pd.DataFrame, key: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    if key not in frame.columns:
        return frame.copy(), pd.DataFrame(columns=list(frame.columns) + ["重复组"])

    normalized = frame[key].map(_normalize_key)
    valid = normalized != ""
    duplicated = valid & normalized.duplicated(keep=False)
    duplicate_rows = frame.loc[duplicated].copy()
    if duplicate_rows.empty:
        return frame.copy(), pd.DataFrame(columns=list(frame.columns) + ["重复组"])

    duplicate_rows["重复组"] = normalized.loc[duplicated].values
    data_columns = [col for col in frame.columns if not col.startswith("来源")]
    completeness = frame[data_columns].apply(lambda row: sum(_text(v) != "" for v in row), axis=1)

    keep_indices: list[int] = []
    for _, indices in normalized.loc[valid].groupby(normalized.loc[valid]).groups.items():
        idx_list = list(indices)
        best = max(idx_list, key=lambda idx: (int(completeness.loc[idx]), -idx))
        keep_indices.append(best)

    keep_indices.extend(frame.index[~valid].tolist())
    kept = frame.loc[sorted(set(keep_indices))].reset_index(drop=True)
    return kept, duplicate_rows.reset_index(drop=True)


def _missing_rows(frame: pd.DataFrame) -> pd.DataFrame:
    required = [column for column in ("学号", "姓名") if column in frame.columns]
    if not required:
        return pd.DataFrame(columns=list(frame.columns) + ["缺失字段"])

    mask = frame[required].apply(lambda row: any(_text(value) == "" for value in row), axis=1)
    missing = frame.loc[mask].copy()
    if missing.empty:
        return pd.DataFrame(columns=list(frame.columns) + ["缺失字段"])
    missing["缺失字段"] = missing.apply(
        lambda row: "、".join(column for column in required if _text(row[column]) == ""), axis=1
    )
    return missing


def _category_stats(frame: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for column in ("学院", "专业", "班级", "校区", "状态", "性别"):
        if column not in frame.columns:
            continue
        counts = frame[column].replace("", "（空白）").value_counts(dropna=False)
        for value, count in counts.items():
            rows.append({"统计字段": column, "分类": _text(value) or "（空白）", "数量": int(count)})
    if not rows:
        rows.append({"统计字段": "总计", "分类": "有效记录", "数量": len(frame)})
    return pd.DataFrame(rows)


def _anomaly_rows(frame: pd.DataFrame, file_logs: list[dict[str, Any]]) -> pd.DataFrame:
    anomalies: list[dict[str, Any]] = []

    if "联系电话" in frame.columns:
        for index, value in frame["联系电话"].items():
            digits = re.sub(r"\D", "", _text(value))
            if value and len(digits) not in {7, 8, 10, 11, 12}:
                anomalies.append({"类型": "联系电话格式", "行号": index + 2, "内容": _text(value), "说明": "号码长度可能异常"})

    if "学号" in frame.columns and "姓名" in frame.columns:
        valid = frame.loc[frame["学号"].map(_normalize_key) != "", ["学号", "姓名"]].copy()
        if not valid.empty:
            name_counts = valid.groupby(valid["学号"].map(_normalize_key))["姓名"].nunique()
            conflict_keys = set(name_counts[name_counts > 1].index)
            for index, row in valid.iterrows():
                if _normalize_key(row["学号"]) in conflict_keys:
                    anomalies.append({"类型": "学号姓名冲突", "行号": index + 2, "内容": f"{row['学号']} / {row['姓名']}", "说明": "同一学号对应多个姓名"})

    for log in file_logs:
        if log.get("状态") == "失败":
            anomalies.append({"类型": "文件处理失败", "行号": "-", "内容": f"{log.get('文件名')} / {log.get('工作表')}", "说明": log.get("说明", "")})

    return pd.DataFrame(anomalies, columns=["类型", "行号", "内容", "说明"])


def _style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index in range(1, sheet.max_column + 1):
            max_length = 0
            for row_index in range(1, min(sheet.max_row, 500) + 1):
                value = sheet.cell(row=row_index, column=column_index).value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 36)
        sheet.sheet_view.showGridLines = False
    workbook.save(path)


def process_files(
    uploaded_paths: list[Path],
    working_dir: Path,
    dedupe_key: str = "学号",
) -> ProcessResult:
    working_dir.mkdir(parents=True, exist_ok=True)
    excel_files, file_logs = collect_excel_files(uploaded_paths, working_dir)
    frames: list[pd.DataFrame] = []

    for path in excel_files:
        read_frames, logs = read_single_file(path)
        frames.extend(read_frames)
        file_logs.extend(logs)

    if not frames:
        raise ValueError("没有读取到任何有效表格数据，请检查文件格式和内容")

    all_data = pd.concat(frames, ignore_index=True, sort=False).fillna("")
    original_count = len(all_data)
    merged, duplicates = _most_complete_rows(all_data, dedupe_key) if dedupe_key else (all_data, pd.DataFrame())
    merged.insert(0, "序号", range(1, len(merged) + 1))

    missing = _missing_rows(merged)
    categories = _category_stats(merged)
    anomalies = _anomaly_rows(merged, file_logs)
    file_log_frame = pd.DataFrame(file_logs, columns=["文件名", "工作表", "状态", "有效行数", "说明"])

    output_path = working_dir / "总汇总表.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="全部明细", index=False)
        categories.to_excel(writer, sheet_name="分类统计", index=False)
        duplicates.to_excel(writer, sheet_name="重复数据", index=False)
        missing.to_excel(writer, sheet_name="缺失数据", index=False)
        anomalies.to_excel(writer, sheet_name="异常数据", index=False)
        file_log_frame.to_excel(writer, sheet_name="文件处理记录", index=False)

    _style_workbook(output_path)

    success_files = len({log["文件名"] for log in file_logs if log.get("状态") == "成功"})
    failed_files = len({log["文件名"] for log in file_logs if log.get("状态") == "失败"})
    return ProcessResult(
        output_path=output_path,
        stats={
            "uploaded_files": len(uploaded_paths),
            "detected_tables": len(frames),
            "success_files": success_files,
            "failed_files": failed_files,
            "original_rows": original_count,
            "result_rows": len(merged),
            "duplicate_rows": len(duplicates),
            "missing_rows": len(missing),
            "anomaly_rows": len(anomalies),
        },
        file_logs=file_logs,
    )
